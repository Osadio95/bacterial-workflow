import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

# Define orange color for cell background
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


# Formatting the checkm2 sheet
def format_checkm2(ws):
    for row in range(2, ws.max_row + 1):
        contamination = ws.cell(row=row, column=4).value  # Column D
        completeness = ws.cell(row=row, column=3).value  # Column C
        n50 = ws.cell(row=row, column=8).value  # Column H

        if contamination is not None and contamination > 5:
            ws.cell(row=row, column=4).fill = orange_fill
        if completeness is not None and completeness < 90:
            ws.cell(row=row, column=3).fill = orange_fill
        if n50 is not None and n50 < 50000:
            ws.cell(row=row, column=8).fill = orange_fill


# Formatting the quast sheet
def format_quast(ws):
    for row in range(2, ws.max_row + 1):
        n50 = ws.cell(row=row, column=19).value  # Column S
        total_length = ws.cell(row=row, column=17).value  # Column Q
        contigs = ws.cell(row=row, column=15).value  # Column O

        if n50 is not None and n50 < 50000:
            ws.cell(row=row, column=19).fill = orange_fill
        if total_length is not None and total_length > 6200000:
            ws.cell(row=row, column=17).fill = orange_fill
        if contigs is not None and contigs > 400:
            ws.cell(row=row, column=15).fill = orange_fill

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, '../results.xlsx')

    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return

    try:
        wb = load_workbook(file_path)
        required_sheets = ['checkm2', 'quast', 'identification', 'AMR']
        missing = [s for s in required_sheets if s not in wb.sheetnames]

        if missing:
            print(f"⚠ Missing sheets: {', '.join(missing)}. No formatting applied.")
            return

        format_checkm2(wb['checkm2'])
        format_quast(wb['quast'])

        wb.save(os.path.join(script_dir, '../results.xlsx'))
        print("✔ Formatting completed successfully")

    except InvalidFileException:
        print("❌ Invalid file format")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")

if __name__ == "__main__":
    main()

